#!/usr/bin/env python3
"""Generate TID_catalog_public.xlsx — a sanitized, publishable Traktor TID reference.

Derived from the internal TID_catalog.xlsx but stripped of all third-party content:
  - the verbatim "Function (labels)" text (authored by mapping creators) is REPLACED with
    our own functional descriptions, left blank where we can't describe a TID honestly;
  - the named-mapping "Sources" column is DROPPED;
  - the project-specific "Gated-by Mod" column is DROPPED.
What remains is factual protocol data (control type, interaction, value ranges) plus our own
hardware-test status — facts + own observations, safe for public/DJTechTools upload.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# --- our own functional descriptions (authored from TID_REFERENCE.md, NOT copied labels) ---
DESC = {
    # transport / deck
    100: "Play / Pause", 125: "Sync", 206: "Cue (jump to cue point / start)",
    2350: "Flux mode toggle", 2380: "Beatjump (value = jump-size index, sign = direction)",
    2588: "Toggle deck focus (A / B)", 9: "Deck focus selector",
    102: "Deck channel volume",
    # loops
    200: "Loop In", 201: "Loop Out", 202: "Loop active on / off", 203: "Reloop",
    2192: "Loop set (auto-loop at current size)",
    2317: "Loop size select + set (value = size index)",
    # hotcues
    2328: "Select / set / store hotcue (focused deck; value = slot index)",
    # FX units (deck/unit field = FX unit 0-3)
    321: "FX Unit 1 -> deck assign", 322: "FX Unit 2 -> deck assign",
    338: "FX Unit 3 -> deck assign", 339: "FX Unit 4 -> deck assign",
    362: "Effect slot 1 selector (value = effect list position)",
    363: "Effect slot 2 selector", 364: "Effect slot 3 selector",
    365: "FX dry/wet (= Decay in Pattern Player mode)",
    366: "FX knob 1 (= Volume in Pattern Player mode)",
    367: "FX knob 2 (= Pattern selector in Pattern Player mode)",
    368: "FX knob 3 (= Pitch in Pattern Player mode)",
    369: "FX unit on (= Play / Mute in Pattern Player mode)",
    370: "FX button 1", 371: "FX button 2", 372: "FX button 3",
    # stems (deck+slot field 0-15)
    251: "Stem / remix slot volume", 259: "Stem / remix slot mute (level-preserving)",
    250: "Stem slot filter on", 249: "Stem slot filter amount", 239: "Stem slot FX send",
    # remix / capture
    2002: "Capture deck loop into remix deck (silent)",
    263: "Capture from loop recorder into remix slot",
    # Loop Recorder transport (confirmed: cmdr KnownCommands.cs + decoded DJTT mapping 220)
    280: "Loop Recorder record / arm",
    281: "Loop Recorder size (4 / 8 / 16 / 32 beats)",
    282: "Loop Recorder dry/wet",
    283: "Loop Recorder play / pause",
    284: "Loop Recorder delete",
    287: "Loop Recorder undo / redo",
    # browser / misc
    3200: "Browser list scroll", 3328: "Browser tree select (value = direction)",
    4209: "Browser view toggle", 8194: "Cruise / Auto-DJ on",
    2056: "Audio recorder record", 2301: "Restore FX preset / snapshot",
    17: "Monitor (cue) mix", 402: "Key adjust (value = semitones / 12)",
    2704: "Main output level L (metering output)", 2705: "Main output level R (metering output)",
    304: "EQ low kill", 305: "EQ mid kill", 306: "EQ high kill",
}
# modifiers #1-#8
for n in range(1, 9):
    DESC[2547 + n] = f"Modifier #{n} (set value / use as condition)"
# remix-deck slot cell triggers (slots 1-4 x cells 1-8)
for t in [600,601,602,603,604,605,606,607, 616,617,618,619,620,621,622,623,
          632,633,634,635,636,637,638,639, 648,649,650,651,652,653,654,655]:
    DESC[t] = "Remix deck slot cell trigger"

STATUS_PUBLIC = {
    "PROVEN": "Proven (hardware-tested)",
    "VERIFIED": "Documented",
    "seen-unlabeled": "Observed",
}

wb_in = openpyxl.load_workbook("TID_catalog.xlsx", read_only=True)
rows = list(wb_in["Catalog"].iter_rows(values_only=True))[1:]   # skip header

wb = openpyxl.Workbook()

# --- About sheet ---
about = wb.active
about.title = "About"
about_lines = [
    ("Traktor TID Command Reference", True),
    ("", False),
    ("A factual reference of Traktor controller command IDs (TIDs) and their control", False),
    ("characteristics: control type, interaction mode, the deck/unit field, and accepted", False),
    ("value ranges. These are observations of how the Traktor MIDI protocol behaves.", False),
    ("", False),
    ("Function descriptions and the 'Proven (hardware-tested)' status are the author's own.", False),
    ("Rows left without a description are real command IDs whose exact function is not", False),
    ("confirmed here; their control-type / value data is still provided.", False),
    ("", False),
    ("No third-party mapping text, names, or files are included.", False),
    ("Status: Proven = confirmed on hardware; Documented = function known; Observed = ID seen.", False),
]
for i, (text, bold) in enumerate(about_lines, 1):
    c = about.cell(row=i, column=1, value=text)
    if bold:
        c.font = Font(bold=True, size=14)
about.column_dimensions["A"].width = 90

# --- Reference sheet ---
ws = wb.create_sheet("TID Reference")
headers = ["TID", "Function", "Control Type", "Interaction",
           "Deck/Unit field values", "Set values", "Status"]
ws.append(headers)
hfill = PatternFill("solid", fgColor="305496")
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = hfill
    c.alignment = Alignment(vertical="center")

green = PatternFill("solid", fgColor="C6EFCE")
blue  = PatternFill("solid", fgColor="DDEBF7")
for r in rows:
    tid, _labels, ctypes, inters, deckvals, setvals, _gated, status, _refs, _sources = r
    pub_status = STATUS_PUBLIC.get(status, status)
    ws.append([tid, DESC.get(tid, ""), ctypes, inters, deckvals, setvals, pub_status])
    row_cells = ws[ws.max_row]
    if status == "PROVEN":
        for c in row_cells: c.fill = green
    elif status == "VERIFIED":
        for c in row_cells: c.fill = blue

widths = [8, 48, 22, 26, 24, 30, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"

wb.save("TID_catalog_public.xlsx")
described = sum(1 for r in rows if DESC.get(r[0]))
print(f"Wrote TID_catalog_public.xlsx: {len(rows)} TIDs, {described} with own descriptions.")
print("Columns: TID | Function | Control Type | Interaction | Deck/Unit values | Set values | Status")
print("Dropped: verbatim labels, named Sources, project-specific Gated-by-Mod.")
